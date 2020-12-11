import openpyxl
from openpyxl import formatting, styles
from openpyxl.formatting import rule
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.chart import (
	LineChart,
	Reference,
)
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill


def add_data_to_sheet(wb, sheet, data, start_row, start_column, style):
	
	row_to_edit = start_row
	column_to_edit = start_column

	for row in data:
		
		for column in row:
			
			data_to_write = column 
			if data_to_write != '':
				sheet[get_column_letter(column_to_edit) + str(row_to_edit)] = data_to_write
				if 'Font' in style.keys():
					sheet[get_column_letter(column_to_edit) + str(row_to_edit)].font = style['Font']
				if 'Border' in style.keys():
					sheet[get_column_letter(column_to_edit) + str(row_to_edit)].border = style['Border']	
				if 'Fill' in style.keys():
					sheet[get_column_letter(column_to_edit) + str(row_to_edit)].fill = style['Fill']
				if 'Alignment' in style.keys():
					sheet[get_column_letter(column_to_edit) + str(row_to_edit)].alignment = style['Alignment']
			column_to_edit += 1
			
		column_to_edit = start_column
		row_to_edit += 1
		
def set_number_format(wb, sheet, no_format, start_row, start_column, end_row, end_column):

	for row in range(start_row, end_row):
		for column in range(start_column, end_column):
			column_letter = get_column_letter(column)

			cell = sheet.cell(row, column)
			cell.number_format = no_format #'0.00E+00'
				
def setup_styles():

	sheet_styles = {}

	all_thin_border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),
					top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))

	all_thin_border_grey = Border(left=Side(border_style='thin',color='808080'),right=Side(border_style='thin',color='808080'),
					top=Side(border_style='thin', color='808080'), bottom=Side(border_style='thin', color='808080'))

	all_thick_border = Border(left=Side(border_style='thick',color='FF000000'),right=Side(border_style='thick',color='FF000000'),
					top=Side(border_style='thick', color='FF000000'), bottom=Side(border_style='thick', color='FF000000'))
					
	normal_font = Font(name='Arial', size=8, bold=False, italic=False, vertAlign=None,
				underline='none', strike=False, color='FF000000')

	normal_font_white = Font(name='Arial', size=8, bold=False, italic=False, vertAlign=None,
				underline='none', strike=False, color='FFFFFF')
				
	input_font = Font(name='Arial', size=8, bold=False, italic=False, vertAlign=None,
				underline='none', strike=False, color='0000ff')
				
	grey_fill = PatternFill("solid", fgColor="bfbfbf")
	grey_blue_fill = PatternFill("solid", fgColor="bacfdc")
	
	dark_grey_fill = PatternFill("solid", fgColor="969696")
	
	center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
	# ########### normal style (arial 8, border all round) ########################################
			
	sheet_styles ['Normal'] = {'Font': normal_font, 'Border': all_thin_border}
				
	# ########### Header style (arial 8 bold) ########################################
	
	font = Font(name='Arial', size=8, bold=True, italic=False, vertAlign=None,
				underline='none', strike=False, color='FF000000')
				
	sheet_styles ['Header'] = {'Font': font, 'Border': all_thin_border, 'Alignment': center_alignment}

	# ########### Project Label style (white, arial 8 bold, grey background) ########################################
	
	sheet_styles ['Project_Label'] = {'Font': normal_font_white, 'Fill': dark_grey_fill}
	
	# ########### Tensioner Label style (arial 8 bold, grey background) ########################################
	
	sheet_styles ['Tensioner_Label'] = {'Font': normal_font, 'Border': all_thin_border, 'Fill': grey_fill}

	# ########### Stackup Header style (arial 10 bold grey background) ########################################
	
	font = Font(name='Arial', size=10, bold=True, italic=False, vertAlign=None,
				underline='none', strike=False, color='FF000000')
				
	sheet_styles ['Stackup Header'] = {'Font': font, 'Border': all_thick_border, 'Fill': grey_fill, 'Alignment': center_alignment}

	# ########### input style (arial 8, blue font) ########################################
			
	sheet_styles ['Input'] = {'Font': input_font, 'Border': all_thin_border}

	# ########### Wood Tables ########################################	
	
	# header font
	font = Font(name='Segoe UI', size=9, bold=True, italic=False, vertAlign=None,
				underline='none', strike=False, color='FF000000')
				#
	sheet_styles ['Word Table Header'] = {'Font': font, 'Border': all_thin_border_grey, 'Fill': grey_blue_fill, 'Alignment': center_alignment}
	
	#main table font
	font = Font(name='Segoe UI', size=9, bold=False, italic=False, vertAlign=None,
				underline='none', strike=False, color='FF000000')
	sheet_styles ['Word Table Main'] = {'Font': font, 'Border': all_thin_border_grey, 'Alignment': center_alignment}
	return sheet_styles			



